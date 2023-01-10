import json
import time
import datetime
import requests
import openpyxl

# Load the data from the JSON file, or initialize an empty list if the file doesn't exist
filename = 'data.json'
server_url = "http://localhost:8000"

try:
    with open(filename, 'r') as f:
        data = json.load(f)
except FileNotFoundError:
    data = []

# ID counter
if 'id' in data[0]:
    id_counter = max([obj['id'] for obj in data], default=1)
else:
    id_counter = 1


def add_object():
    global id_counter
    id_counter += 1
    date = None

    if not data:
        # If the data list is empty, prompt the user to enter the keys
        keys = input("Enter keys for new object (comma-separated): ").split(",")
        keys = [key.strip() for key in keys]

        include_date = input("Do you want to include a date with this object? (y/n) ")
        if include_date.lower() == "y":
            now = datetime.datetime.now()
            date = f"{now.day:02d}-{now.month:02d}-{now.year:02d} {now.hour:02d}:{now.minute:02d}"
    else:
        # If the data list is not empty, use the keys from the first object
        keys = list(data[0].keys())
        if 'date' in data[0]:
            date = data[0]['date']

    # Create a new dictionary for the object
    new_object = {
        'id': id_counter,
    }

    if date is not None:
        new_object['date'] = date

    # Prompt the user to enter values for each key
    for key in keys:
        if key == "id" or key == "date":
            continue
        value = input(f"{key}: ")
        if data and key in data[0]:
            # If the data list is not empty and the key is present in the first object, use the same type as the
            # value in the first object. If value type is different from the original object, try to find suitable type
            value_type = type(data[0][key])
            try:
                if value_type == int:
                    value = int(value)
                elif value_type == float:
                    value = float(value)
                elif value_type == str:
                    value = str(value)
            except:
                try:
                    value = int(value)
                except:
                    try:
                        value = float(value)
                    except:
                        value = str(value)
        else:
            try:
                value = int(value)
            except:
                try:
                    value = float(value)
                except:
                    value = str(value)

        new_object[key] = value

    # Append the new object to the data list
    data.append(new_object)
    assign_id()
    check_ids()
    save_data()

    for obj in data:
        id_number = obj.pop('id')
        obj.update({'id': id_number})

    print(f"Object added {max([obj['id'] for obj in data], default=1)}")


def search_objects():
    search_term = input("Enter search term: ")
    search_results = []
    for obj in data:
        if search_term in str(obj.values()):
            search_results.append(obj)
    if search_results:
        print("Search results:")
        for obj in search_results:
            print(obj)
    else:
        print("No results found")


def delete_object():
    id_to_delete = input("Enter ID(s) of object(s) to delete (comma-separated or 'all'): ")
    if id_to_delete == 'all':
        confirm = input("Are you sure you want to delete all objects? (y/n) ")
        if confirm == 'y':
            data.clear()
            print("All objects deleted")
            check_ids()
            save_data()
        else:
            print("Deletion cancelled")
        return
    ids_to_delete = [int(x) for x in id_to_delete.split(',')]
    delete_count = 0
    not_found = 0
    for id_to_delete in ids_to_delete:
        for i, obj in enumerate(data):
            if obj['id'] == id_to_delete:
                del data[i]
                delete_count += 1
                shift_ids(id_to_delete)
                print(f"Object with ID {id_to_delete} deleted")
                break
        else:
            not_found += 1
            print(f"Object with ID {id_to_delete} not found")
    if delete_count or not_found:
        check_ids()
        save_data()


def shift_ids(id_to_delete):
    for obj in data:
        if obj['id'] > id_to_delete:
            obj['id'] -= 1


def check_ids():
    for i, obj in enumerate(data):
        if obj['id'] != i + 1:
            obj['id'] = i + 1


def assign_id():
    if 'id' not in data[0]:
        for i, obj in enumerate(data):
            obj["id"] = i


def sort_data():
    keys = list(data[0].keys())
    print("Enter key to sort by:")
    for i, key in enumerate(keys):
        print(f"{i + 1}. {key}")
    choice = int(input("> "))
    key = keys[choice - 1]
    data.sort(key=lambda x: x[key])


def list_all_data():
    print("All data:")
    for obj in data:
        print(obj)


def save_data():
    with open(filename, 'w') as f:
        json.dump(data, f, indent=4)


def edit_object():
    id_to_edit = input("Enter ID of object to edit: ")
    for obj in data:
        if obj['id'] == int(id_to_edit):
            print("Enter new values or leave blank to keep current value")
            for key in obj.keys():
                if key == 'id':
                    continue
                new_value = input(f"{key}: ")
                if new_value:
                    # Check if the new value is a number
                    if new_value.isdigit():
                        # If the original value is a float, update it with a float
                        if isinstance(obj[key], float):
                            obj[key] = float(new_value)
                        # Otherwise, update it with an integer
                        else:
                            obj[key] = int(new_value)
                    # If the new value is not a number, update it with a string
                    else:
                        obj[key] = new_value
            save_data()
            print("Object updated")
            return
    print("Object not found")


def sync_data():
    # Make a GET request to the server to get the data from the server
    response = requests.get(server_url)

    # Check the status code of the response
    if response.status_code == 200:
        # If the request is successful, get the data from the response
        server_data = response.json()

        # Open the local JSON file
        with open("data.json", "r") as f:
            # Read the contents of the file
            file_data = json.load(f)

        # Compare the data in the server with the data in the file
        if server_data != file_data:
            # If the data is different, ask the user if they want to overwrite the local data with the data from the
            # server
            answer = input("The data in the server is different from the data in the file. Do you want to overwrite "
                           "the local data with the data from the server? (y/n) ")
            if answer.lower() == "y":
                # If the user wants to overwrite the local data, make a POST request to the server to update the data
                response = requests.post(server_url, json=server_data)

                # Check the status code of the response
                if response.status_code == 200:
                    # If the request is successful, save the data from the server to the local file
                    with open("data.json", "w") as f:
                        json.dump(server_data, f)
                else:
                    print("Error updating data:", response.status_code)
            elif answer.lower() == "n":
                # If the user does not want to overwrite the local data, ask if they want to update the server with
                # the local data
                answer = input("Do you want to update the server with the local data? (y/n) ")
                if answer.lower() == "y":
                    # If the user wants to update the server with the local data, make a POST request to the server
                    # to update the data
                    response = requests.post(server_url, json=file_data)

                    # Check the status code of the response
                    if response.status_code == 200:
                        print("Data updated successfully")
                    else:
                        print("Error updating data:", response.status_code)
                else:
                    print("Data not updated")
        else:
            print("Data is already up to date")
    else:
        print("Error getting data from server:", response.status_code)


def convert_to_excel():
    # Load the data from the JSON file
    with open(filename, "r") as f:
        data = json.load(f)

    # If file opened successfully, convert it, if not - print the message
    if data:
        # Create a new workbook
        workbook = openpyxl.Workbook()
        # Get the active worksheet
        worksheet = workbook.active

        # Get the keys from the first object in the data list
        keys = list(data[0].keys())
        # Write the keys to the worksheet as the column headers
        worksheet.append(keys)

        # Iterate over the data and write each row to the worksheet
        for row in data:
            worksheet.append([row[key] for key in keys])

        # Prompt the user for the name of the Excel file to save
        excel_filename = input("Enter the name of the Excel file to save: ")
        # Save the workbook
        workbook.save(excel_filename + ".xlsx")

        # Print a message indicating that the conversion was successful
        print(f"Successfully converted {len(data)} rows of data from '{filename}' to '{excel_filename}'.")
    else:
        print("This file is empty")


def convert_to_json():
    # Get the name of the Excel file to be converted
    excel_filename = input("Enter the name of the Excel file to be converted: ")
    # Get the name of the JSON file to update
    json_filename = input("Enter the name of the JSON file to update: ")

    # Load the workbook from the Excel file
    workbook = openpyxl.load_workbook(excel_filename)
    # Get the active worksheet
    worksheet = workbook.active

    # Get the names of the columns from the first row of the worksheet
    columns = [cell.value for cell in worksheet[1]]

    # Create a new list to store the data
    data = []
    # Iterate over the rows of the worksheet, starting from the second row (row 1)
    for row in worksheet.iter_rows(min_row=2):
        # Create a new dictionary to store the data for this row
        row_data = {}
        # Iterate over the cells in the row
        for cell, column in zip(row, columns):
            # Add the data from the cell to the row_data dictionary
            row_data[column] = cell.value
        # Add the row_data dictionary to the data list
        data.append(row_data)

    # Open the JSON file in write mode
    with open(json_filename, "w") as f:
        # Write the data to the JSON file, overwriting its contents
        json.dump(data, f, indent=4, default=str)
    print(f"Successfully converted {len(data)} rows of data from '{excel_filename}' to '{json_filename}'.")


# Main loop
while True:
    print("\nEnter a command:")
    print("1. Add object")
    print("2. Search objects")
    print("3. Delete object")
    print("4. Sort data")
    print("5. List all data")
    print("6. Edit object")
    print("7. Get data from server")
    print("8. Convert file to Excel")
    print("9. Convert Excel file to json")
    print("0. Quit")
    command = input("> ")
    if command == '1':
        add_object()
    elif command == '2':
        search_objects()
    elif command == '3':
        delete_object()
    elif command == '4':
        sort_data()
    elif command == '5':
        list_all_data()
    elif command == '6':
        edit_object()
    elif command == '7':
        sync_data()
    elif command == '8':
        convert_to_excel()
    elif command == '9':
        convert_to_json()
        with open("data.json", "r") as f:
            data = json.load(f)
    elif command == '0':
        break
    else:
        print("Invalid command")

# Save the data to the JSON file
save_data()
